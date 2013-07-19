#!/usr/bin/env python
# -*- coding: utf-8 -*-

import requests # requires Python requests 1.x
import optparse
from openpyxl.workbook import Workbook
from openpyxl.writer.excel import ExcelWriter
import unicodecsv
import json
import collections
import scraperwiki
import datetime
from tempfile import mkstemp
import os
import sys
from os.path import join, abspath, dirname

DEBUG = True # prints debug messages to stdout during run

MAX_ROWS = 5000 # how many rows to request from the SQL API at any one time

USAGE = """Convert data from a ScraperWiki box into CSVs and Excel spreadsheets.
Reads from a file in the home directory containing  the full URL of the target box,
including publishToken, ie http://box.scraperwiki.com/boxName/publishToken"""
def main():
    try:
        filename = abspath(join(dirname(__file__), '..', 'dataset_url.txt'))
        with open(filename, 'r') as f:
            box_url = f.read().strip()
    except IOError:
        print("ERROR: No dataset URL in {}, try hitting regenerate.\n".format(filename))
        print(USAGE)
        sys.exit(1)

    scraperwiki.sql.execute('CREATE TABLE IF NOT EXISTS "_state" ("filename" UNIQUE, "created")')
    scraperwiki.sql.commit()

    tables_and_columns = get_tables_and_columns(box_url)
    log(tables_and_columns)

    # This might look a bit complicated, because we're creating
    # a multi-sheet XLSX and a bunch of CSV files at the same time.
    # But it's more efficient than two separate loops.
    # We save state into the database, for the GUI to read.
    excel_workbook = Workbook(optimized_write = True)
    save_state('all_tables.xlsx', 'creating')
    for table_name, column_names in tables_and_columns.items():
        csv_tempfile = make_temp_file('.csv')
        save_state("%s.csv" % table_name, 'creating')
        with open(csv_tempfile, 'wb') as f:
            # NOTE: create_sheet(title=foo) doesn't appear to name the sheet in
            # openpyxl version 1.5.7, hence manually setting title afterwards.
            excel_worksheet = excel_workbook.create_sheet(title=table_name)
            excel_worksheet.title = table_name
            excel_worksheet.append(column_names)
            csv_writer = unicodecsv.DictWriter(f, column_names)
            csv_writer.writeheader()
            for chunk_of_rows in get_rows(box_url, table_name):
                csv_writer.writerows(chunk_of_rows)
                for row in chunk_of_rows:
                    excel_worksheet.append(row.values())
        replace_tempfile(csv_tempfile, "http/%s.csv" % table_name)
        save_state("%s.csv" % table_name, 'completed')

    xlsx_tempfile = make_temp_file('.xlsx')
    excel_workbook.save(filename=xlsx_tempfile)
    replace_tempfile(xlsx_tempfile, 'http/all_tables.xlsx')
    save_state('all_tables.xlsx', 'completed')


def make_temp_file(suffix):
    (_, filename) = mkstemp(suffix=suffix, dir='http')
    os.chmod(filename, 0644)  # world-readable
    return filename


def replace_tempfile(tmp, destination):
    os.rename(tmp, destination)


def log(string):
    if DEBUG: print string

def call_api(box_url, params=None):
    # returns sql api output as a Python dict/list
    log("call_api(%s)" % box_url)
    response = requests.get(box_url, params=params)
    log("GET %s" % response.url)
    if response.status_code == requests.codes.ok:
        return json.loads(response.content, object_pairs_hook=collections.OrderedDict)
    else:
        response.raise_for_status()

def query_sql_database(box_url, query):
    return call_api("%s/sql" % box_url, {"q": query})

def get_database_meta(box_url):
    return call_api("%s/sql/meta" % box_url)

def get_tables_and_columns(box_url):
    # returns a dict of lists, like so:
    # { tableOne: [ col1, col2 ], tableTwo: […, …] }
    meta = get_database_meta(box_url)
    result = {}
    for table_name in meta['table'].keys():
        result[table_name] = meta['table'][table_name]['columnNames']
    return result

def get_rows(box_url, table_name):
    start = 0
    while True:
        rows = query_sql_database(box_url, """SELECT * FROM "%s" LIMIT %d, %d""" % (table_name, start, MAX_ROWS))
        if not rows:
            break
        yield rows
        start += MAX_ROWS

def save_state(filename, state):
    log("%s %s" % (filename, state))
    if state == 'creating':
        scraperwiki.sql.save(['filename'], {'filename': filename, "created": None}, '_state')
    elif state == 'completed':
        now = scraperwiki.sql.select('datetime("now") as now')[0]['now']
        scraperwiki.sql.save(['filename'], {'filename': filename, "created": now}, '_state')
    else:
        raise Exception("Unknown status: %s" % state)

try:
    main()
except Exception as e:
    print '"We encountered a Python error while extracting your dataset: %s"' % e
    if DEBUG: raise
